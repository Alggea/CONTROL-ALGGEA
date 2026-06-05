"""Export utilities: render Excel (.xlsx) and PDF reports from row data."""
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

HEADER_FILL = PatternFill(start_color="002FA7", end_color="002FA7", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="E4E4E7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def to_xlsx(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    # Title row
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="002FA7")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    # Subtitle (date)
    ws.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color="64748B")
    ws.append([])
    # Header
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = CELL_BORDER
    # Rows
    for r in rows:
        ws.append(r)
    # Borders + auto width
    for row in ws.iter_rows(min_row=5, max_row=4 + len(rows), max_col=len(headers)):
        for cell in row:
            cell.border = CELL_BORDER
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    for col_idx, header in enumerate(headers, start=1):
        letter = ws.cell(row=4, column=col_idx).column_letter
        max_len = max(
            [len(str(header))] + [len(str(r[col_idx - 1])) if col_idx - 1 < len(r) else 0 for r in rows]
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)
    ws.row_dimensions[4].height = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(title: str, headers: List[str], rows: List[List[Any]]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"],
        textColor=colors.HexColor("#002FA7"),
        fontSize=18, leading=22, spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"],
        textColor=colors.HexColor("#64748B"),
        fontSize=9, leading=12, spaceAfter=12,
    )
    elements = [
        Paragraph(title, title_style),
        Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} &nbsp;·&nbsp; {len(rows)} registros",
            meta_style,
        ),
    ]
    data = [headers] + [[str(c) if c is not None else "" for c in row] for row in rows]
    if not rows:
        data.append(["(Sin datos)"] + [""] * (len(headers) - 1))
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002FA7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
        ("TOPPADDING", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E4E4E7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(tbl)
    doc.build(elements)
    return buf.getvalue()


def fmt_currency(n) -> str:
    try:
        return f"${float(n):,.2f}"
    except Exception:
        return str(n)


def fmt_date_short(s) -> str:
    if not s:
        return ""
    try:
        d = datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        return d.strftime("%d/%m/%Y")
    except Exception:
        return str(s)[:10]
