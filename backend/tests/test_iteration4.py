"""Iteration 4 tests: Partner Portal banner, Projects filters (q/client_id), Exports (xlsx/pdf)."""
import os
import io
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
EMAIL = "carlos@socios.com"
PASSWORD = "socio123"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=20)
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def hdr(token):
    return {"Authorization": f"Bearer {token}"}


# -------------------- Partner Portal banner --------------------
class TestPartnerPortal:
    def test_portal_includes_new_fields(self, hdr):
        r = requests.get(f"{BASE_URL}/api/partners/portal", headers=hdr, timeout=20)
        assert r.status_code == 200
        data = r.json()
        for k in ("net_balance", "per_partner_share", "total_dividends_withdrawn",
                  "total_reimbursements_paid", "available_to_distribute", "partners"):
            assert k in data, f"missing field {k}"
        # math: available = net_balance - total_dividends_withdrawn
        assert abs(data["available_to_distribute"] - (data["net_balance"] - data["total_dividends_withdrawn"])) < 0.01
        assert isinstance(data["partners"], list) and len(data["partners"]) >= 1


# -------------------- Projects filters --------------------
class TestProjectsFilters:
    def test_projects_filter_by_q(self, hdr):
        # seed: create a project with a unique name
        unique = "TEST_iter4_alpha_proj"
        cr = requests.post(f"{BASE_URL}/api/projects", headers=hdr, json={
            "name": unique, "status": "in_progress", "start_date": "2025-01-01"
        }, timeout=10)
        assert cr.status_code == 200, cr.text
        pid = cr.json()["id"]
        try:
            r = requests.get(f"{BASE_URL}/api/projects?q=alpha_proj", headers=hdr, timeout=10)
            assert r.status_code == 200
            names = [p["name"] for p in r.json()]
            assert unique in names, f"q filter did not return seeded project; got names={names[:5]}"
            # strict: every returned project must match the q substring (case-insensitive)
            non_matching = [n for n in names if "alpha_proj" not in n.lower()]
            assert not non_matching, f"q filter NOT applied - returned non-matching projects: {non_matching[:5]}"
            # case-insensitive
            r2 = requests.get(f"{BASE_URL}/api/projects?q=ALPHA_PROJ", headers=hdr, timeout=10)
            assert r2.status_code == 200
            assert any(p["name"] == unique for p in r2.json())
        finally:
            requests.delete(f"{BASE_URL}/api/projects/{pid}", headers=hdr, timeout=10)

    def test_projects_filter_by_client_id(self, hdr):
        # create a client + project linked to it
        cl = requests.post(f"{BASE_URL}/api/clients", headers=hdr, json={"name": "TEST_iter4_client"}, timeout=10)
        assert cl.status_code == 200
        cid = cl.json()["id"]
        pr = requests.post(f"{BASE_URL}/api/projects", headers=hdr, json={
            "name": "TEST_iter4_clientproj", "client_id": cid, "status": "in_progress", "start_date": "2025-01-01"
        }, timeout=10)
        assert pr.status_code == 200
        pid = pr.json()["id"]
        try:
            r = requests.get(f"{BASE_URL}/api/projects?client_id={cid}", headers=hdr, timeout=10)
            assert r.status_code == 200
            ids = [p["id"] for p in r.json()]
            assert pid in ids, f"client_id filter did not return project. ids={ids}"
            # make sure non-matching client_id projects are excluded
            for p in r.json():
                assert p.get("client_id") == cid
        finally:
            requests.delete(f"{BASE_URL}/api/projects/{pid}", headers=hdr, timeout=10)
            requests.delete(f"{BASE_URL}/api/clients/{cid}", headers=hdr, timeout=10)

    def test_projects_combine_filters(self, hdr):
        cl = requests.post(f"{BASE_URL}/api/clients", headers=hdr, json={"name": "TEST_iter4_combo_client"}, timeout=10)
        cid = cl.json()["id"]
        p1 = requests.post(f"{BASE_URL}/api/projects", headers=hdr, json={
            "name": "TEST_iter4_combo_inProg", "client_id": cid, "status": "in_progress", "start_date": "2025-01-01"
        }, timeout=10).json()
        p2 = requests.post(f"{BASE_URL}/api/projects", headers=hdr, json={
            "name": "TEST_iter4_combo_done", "client_id": cid, "status": "completed", "start_date": "2025-01-01"
        }, timeout=10).json()
        try:
            r = requests.get(f"{BASE_URL}/api/projects?client_id={cid}&status=in_progress&q=combo", headers=hdr, timeout=10)
            assert r.status_code == 200
            ids = [p["id"] for p in r.json()]
            assert p1["id"] in ids
            assert p2["id"] not in ids
        finally:
            requests.delete(f"{BASE_URL}/api/projects/{p1['id']}", headers=hdr, timeout=10)
            requests.delete(f"{BASE_URL}/api/projects/{p2['id']}", headers=hdr, timeout=10)
            requests.delete(f"{BASE_URL}/api/clients/{cid}", headers=hdr, timeout=10)


# -------------------- Exports --------------------
XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _assert_xlsx(resp, expected_sheet_substr=None):
    assert resp.status_code == 200, resp.text[:300]
    assert XLSX_CT in resp.headers.get("Content-Type", "")
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 4  # title + meta + blank + header
    if expected_sheet_substr:
        assert expected_sheet_substr.lower() in ws.title.lower()


def _assert_pdf(resp):
    assert resp.status_code == 200
    assert resp.headers.get("Content-Type", "").startswith("application/pdf")
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    assert resp.content[:4] == b"%PDF"


class TestExports:
    def test_transactions_xlsx(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/transactions?format=xlsx", headers=hdr, timeout=30)
        _assert_xlsx(r)

    def test_transactions_pdf(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/transactions?format=pdf", headers=hdr, timeout=30)
        _assert_pdf(r)

    def test_transactions_xlsx_with_filters(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/transactions?format=xlsx&type=income&payment_method=transfer", headers=hdr, timeout=30)
        _assert_xlsx(r)

    def test_projects_xlsx_with_q_and_status(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/projects?format=xlsx&q=test&status=in_progress", headers=hdr, timeout=30)
        _assert_xlsx(r)

    def test_projects_pdf(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/projects?format=pdf", headers=hdr, timeout=30)
        _assert_pdf(r)

    def test_clients_xlsx_and_pdf(self, hdr):
        rx = requests.get(f"{BASE_URL}/api/exports/clients?format=xlsx", headers=hdr, timeout=30)
        _assert_xlsx(rx)
        rp = requests.get(f"{BASE_URL}/api/exports/clients?format=pdf", headers=hdr, timeout=30)
        _assert_pdf(rp)

    def test_providers_xlsx_and_pdf(self, hdr):
        rx = requests.get(f"{BASE_URL}/api/exports/providers?format=xlsx", headers=hdr, timeout=30)
        _assert_xlsx(rx)
        rp = requests.get(f"{BASE_URL}/api/exports/providers?format=pdf", headers=hdr, timeout=30)
        _assert_pdf(rp)

    def test_audit_xlsx_filtered(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/audit?format=xlsx&entity_type=transaction", headers=hdr, timeout=30)
        _assert_xlsx(r)

    def test_export_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/exports/transactions?format=xlsx", timeout=10)
        assert r.status_code == 401

    def test_export_invalid_format(self, hdr):
        r = requests.get(f"{BASE_URL}/api/exports/transactions?format=csv", headers=hdr, timeout=10)
        assert r.status_code == 422
